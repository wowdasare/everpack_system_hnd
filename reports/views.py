from django.shortcuts import render
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, Q
from datetime import datetime, timedelta
from sales.models import Sale, Customer, Payment
from inventory.models import Product, StockMovement
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import io


class ReportsHomeView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        current_month = today.replace(day=1)
        
        # Quick statistics
        context['total_sales_this_month'] = Sale.objects.filter(
            sale_date__date__gte=current_month
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        context['products_in_stock'] = Product.objects.filter(
            is_active=True
        ).count()
        
        context['active_customers'] = Customer.objects.filter(
            is_active=True
        ).count()
        
        context['low_stock_items'] = Product.objects.filter(
            is_active=True
        ).count()  # This should be calculated properly
        
        context['overdue_payments'] = Sale.objects.filter(
            payment_status='OVERDUE'
        ).count()
        
        return context


class SalesReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/sales_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Date range filters
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # Sales data
        context['total_sales'] = Sale.objects.count()
        context['total_revenue'] = Sale.objects.aggregate(
            total=Sum('total_amount'))['total'] or 0
        context['avg_order_value'] = Sale.objects.aggregate(
            avg=Avg('total_amount'))['avg'] or 0
        
        context['today_sales'] = Sale.objects.filter(
            sale_date__date=today
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        context['week_sales'] = Sale.objects.filter(
            sale_date__date__gte=week_ago
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        context['month_sales'] = Sale.objects.filter(
            sale_date__date__gte=month_ago
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Recent sales
        context['recent_sales'] = Sale.objects.select_related('customer').order_by('-created_at')[:10]
        
        # Payment status breakdown
        context['paid_sales'] = Sale.objects.filter(payment_status='PAID').count()
        context['pending_sales'] = Sale.objects.filter(payment_status='PENDING').count()
        context['partial_sales'] = Sale.objects.filter(payment_status='PARTIAL').count()
        
        return context


class InventoryReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/inventory_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Inventory statistics
        context['total_products'] = Product.objects.filter(is_active=True).count()
        context['total_stock_value'] = sum(p.stock_value for p in Product.objects.filter(is_active=True))
        
        # Low stock products
        context['low_stock_products'] = [p for p in Product.objects.filter(is_active=True) if p.is_low_stock]
        
        # Recent stock movements
        context['recent_movements'] = StockMovement.objects.select_related('product', 'created_by').order_by('-created_at')[:15]
        
        # Category breakdown
        context['product_categories'] = Product.objects.filter(is_active=True).values(
            'category__name'
        ).annotate(count=Count('id')).order_by('-count')
        
        return context


class FinancialReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/financial_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Financial metrics
        context['total_revenue'] = Sale.objects.aggregate(total=Sum('total_amount'))['total'] or 0
        context['outstanding_amount'] = Sale.objects.filter(
            payment_status__in=['PENDING', 'PARTIAL']
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Profit calculation (simplified)
        total_profit = 0
        for sale in Sale.objects.all():
            total_profit += sale.total_profit
        context['total_profit'] = total_profit
        
        return context


class CustomerReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/customer_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Customer statistics
        context['total_customers'] = Customer.objects.count()
        context['active_customers'] = Customer.objects.filter(is_active=True).count()
        
        # Top customers by purchase amount
        context['top_customers'] = Customer.objects.annotate(
            total_purchases=Sum('sales__total_amount')
        ).filter(total_purchases__gt=0).order_by('-total_purchases')[:10]
        
        # Customer type breakdown
        context['customer_types'] = Customer.objects.values('customer_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return context


class ProfitLossReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/profit_loss_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate profit/loss
        total_revenue = Sale.objects.aggregate(total=Sum('total_amount'))['total'] or 0
        total_cost = sum(p.cost_price * p.current_stock for p in Product.objects.filter(is_active=True))
        
        context['total_revenue'] = total_revenue
        context['total_cost'] = total_cost
        context['gross_profit'] = total_revenue - total_cost
        
        return context


def export_sales_excel(request):
    """Export sales report to Excel format"""
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = 'EverPack System - Sales Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Date
    ws['A2'] = f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}'
    ws.merge_cells('A2:F2')
    
    # Headers
    headers = ['Invoice Number', 'Customer', 'Sale Date', 'Total Amount (GHS)', 'Payment Status', 'Profit (GHS)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data
    sales = Sale.objects.select_related('customer').order_by('-sale_date')
    row = 5
    total_amount = 0
    total_profit = 0
    
    for sale in sales:
        ws.cell(row=row, column=1, value=sale.invoice_number)
        ws.cell(row=row, column=2, value=sale.customer.name)
        ws.cell(row=row, column=3, value=sale.sale_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=4, value=float(sale.total_amount))
        ws.cell(row=row, column=5, value=sale.get_payment_status_display())
        ws.cell(row=row, column=6, value=float(sale.total_profit))
        
        total_amount += sale.total_amount
        total_profit += sale.total_profit
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=3, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_amount)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_profit)).font = Font(bold=True)
    
    # Adjust column widths for sales export
    column_widths = [20, 25, 15, 18, 18, 15]  # Predefined widths
    for i, width in enumerate(column_widths, 1):
        column_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


def export_inventory_excel(request):
    """Export inventory report to Excel format"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = 'EverPack System - Inventory Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Date
    ws['A2'] = f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}'
    ws.merge_cells('A2:H2')
    
    # Headers
    headers = ['SKU', 'Product Name', 'Category', 'Current Stock', 'Cost Price (GHS)', 'Selling Price (GHS)', 'Stock Value (GHS)', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data
    products = Product.objects.select_related('category').filter(is_active=True).order_by('name')
    row = 5
    total_stock_value = 0
    
    for product in products:
        ws.cell(row=row, column=1, value=product.sku)
        ws.cell(row=row, column=2, value=product.name)
        ws.cell(row=row, column=3, value=product.category.name if product.category else 'N/A')
        ws.cell(row=row, column=4, value=product.current_stock)
        ws.cell(row=row, column=5, value=float(product.cost_price))
        ws.cell(row=row, column=6, value=float(product.selling_price))
        ws.cell(row=row, column=7, value=float(product.stock_value))
        ws.cell(row=row, column=8, value='Low Stock' if product.is_low_stock else 'Normal')
        
        total_stock_value += product.stock_value
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=6, value="TOTAL STOCK VALUE:").font = Font(bold=True)
    ws.cell(row=row, column=7, value=float(total_stock_value)).font = Font(bold=True)
    
    # Adjust column widths for inventory export
    column_widths = [15, 30, 15, 12, 15, 15, 18, 12]  # Predefined widths for all 8 columns
    for i, width in enumerate(column_widths, 1):
        column_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


def export_sales_pdf(request):
    """Export sales report to PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("EverPack System - Sales Report", title_style)
    elements.append(title)
    
    # Date
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        spaceAfter=20
    )
    date_para = Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style)
    elements.append(date_para)
    
    # Table data
    data = [['Invoice Number', 'Customer', 'Date', 'Amount (GHS)', 'Status', 'Profit (GHS)']]
    
    sales = Sale.objects.select_related('customer').order_by('-sale_date')
    total_amount = 0
    total_profit = 0
    
    for sale in sales:
        data.append([
            sale.invoice_number,
            sale.customer.name,
            sale.sale_date.strftime("%Y-%m-%d"),
            f"{sale.total_amount:.2f}",
            sale.get_payment_status_display(),
            f"{sale.total_profit:.2f}"
        ])
        total_amount += sale.total_amount
        total_profit += sale.total_profit
    
    # Add totals row
    data.append(['', '', 'TOTAL:', f"{total_amount:.2f}", '', f"{total_profit:.2f}"])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF data and create response
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response


def export_inventory_pdf(request):
    """Export inventory report to PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("EverPack System - Inventory Report", title_style)
    elements.append(title)
    
    # Date
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        spaceAfter=20
    )
    date_para = Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style)
    elements.append(date_para)
    
    # Table data
    data = [['SKU', 'Product Name', 'Stock', 'Cost (GHS)', 'Price (GHS)', 'Value (GHS)', 'Status']]
    
    products = Product.objects.select_related('category').filter(is_active=True).order_by('name')
    total_stock_value = 0
    
    for product in products:
        data.append([
            product.sku,
            product.name[:25] + '...' if len(product.name) > 25 else product.name,
            str(product.current_stock),
            f"{product.cost_price:.2f}",
            f"{product.selling_price:.2f}",
            f"{product.stock_value:.2f}",
            'Low Stock' if product.is_low_stock else 'Normal'
        ])
        total_stock_value += product.stock_value
    
    # Add totals row
    data.append(['', '', '', '', 'TOTAL:', f"{total_stock_value:.2f}", ''])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF data and create response
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response
